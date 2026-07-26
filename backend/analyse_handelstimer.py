#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
analyse_handelstimer.py
=======================
Beskrivende bevaegelses-statistik for MES (Micro E-mini S&P 500) pr. time-paa-
doegnet i DANSK tid — ud fra ~2 aars 1-minuts candles.

Formaal (jf. spec): find hvilke timer paa doegnet der har nok bevaegelse til at
det kan betale sig at handle, i BEGGE retninger. INGEN strategi, ingen entry/
exit, ingen P&L — kun ren beskrivende statistik: hvor meget flytter prisen sig i
denne time, og hvor paalideligt?

Kilde-data: data_harvest/mes_m2k_stitched/MES_1min.csv
  - Kolonner: timestamp, open, high, low, close, volume
  - Tidsstempler er TIDSZONE-BEVIDSTE med eksplicit offset (-04:00 om sommeren /
    -05:00 om vinteren) => America/New_York (ET). Fordi offset er eksplicit, er
    hver raekke et entydigt absolut oejeblik; tz_convert til Europe/Copenhagen er
    derfor eksakt (haandterer dansk sommertid CET<->CEST automatisk).

Output (skrives til OUT_DIR):
  - handelstimer_dansktid.csv   (hoveddeliverable: 24 timer)
  - handelstimer_ugedag.csv     (ugedag x time, langt format)
  - handelstimer_rapport.md     (tabeller + korte konklusioner + split-half)
Plus en paen tabel i terminalen.

Koeres: python analyse_handelstimer.py
"""
from __future__ import annotations

import sys
from pathlib import Path
import pandas as pd
import numpy as np

# Windows-terminal er cp1252 og kan ikke printe emojis/em-dash — tving UTF-8.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# =============================================================================
# KONSTANTER — juster her, ikke i logikken nedenfor
# =============================================================================
HERE       = Path(__file__).resolve().parent
DATA_FILE  = HERE / "data_harvest" / "mes_m2k_stitched" / "MES_1min.csv"
OUT_DIR    = HERE / "analyse_handelstimer_out"

SRC_TZ_NOTE = "Kilde er tz-bevidst ET (offset i strengen) — parses med utc=True"
TARGET_TZ   = "Europe/Copenhagen"     # ALT output er i dansk vaegur-tid

# --- Enheder --------------------------------------------------------------
POINT_PER_TICK = 0.25                  # MES: 1 tick = 0.25 indekspoint
USD_PER_POINT  = 5.0                   # MES: $5 pr. indekspoint

# --- Data-kvalitet --------------------------------------------------------
MIN_MIN_PER_HOUR = 30                  # <30 minutter i en time => tynd/afbrudt time (markeres, forurener ikke)
MIN_N            = 30                  # <30 observationsdage for en time/celle => "for lidt data" (⚪)

# --- "Store timer" (konsistens-maal) --------------------------------------
# En times range regnes "stor" hvis den overstiger DAGENS EGEN median-range paa
# tvaers af doegnets timer (adaptiv, dagsspecifik skaering). Saet til et fast
# point-tal i stedet ved at saette BIG_HOUR_FIXED_PT (None = brug dags-median).
BIG_HOUR_FIXED_PT = None
CONSISTENCY_MIN   = 0.40               # andel store timer under dette => nedgrader ét trin

# --- Retnings-bias --------------------------------------------------------
BIAS_DEADZONE_PT = 0.5                 # |median netto-afkast| under dette => neutral (ingen bias)

# --- Absolutte trafiklys-graenser (median range i point) ------------------
ABS_GREEN_PT  = 5.0                    # >= 5 pt  => absolut groen-kandidat
ABS_YELLOW_PT = 3.0                    # 3-5 pt   => gul; < 3 pt => roed

# --- Median/gennemsnit-divergens -----------------------------------------
DIVERGENCE_FACTOR = 1.5                # gns_range > median_range * dette => flag (faa ekstreme dage traekker)

# Danske ugedagsnavne (0=mandag ... 6=soendag)
DK_DOW = ["Mandag", "Tirsdag", "Onsdag", "Torsdag", "Fredag", "Loerdag", "Soendag"]

# De seks maal vi beregner pr. time-bar (kolonnenavn -> visningsnavn)
MEASURES = {
    "range":    "Range",
    "abs_move": "|beveg.|",
    "net":      "Netto-afkast",
    "max_up":   "Max op",
    "max_down": "Max ned",
    "churn":    "Churn/path",
}


# =============================================================================
# 0-1. INDLAES + TIDSZONE-KONVERTERING
# =============================================================================
def load_minute_data() -> pd.DataFrame:
    """Laes 1-min CSV, konverter til dansk tid, tilfoej danske vaegur-felter."""
    df = pd.read_csv(DATA_FILE)
    # Tz-bevidst parse: utc=True samler de blandede offsets (-04/-05) til eet
    # entydigt UTC-instant pr. raekke, hvorefter vi konverterer til dansk tid.
    ts_utc = pd.to_datetime(df["timestamp"], utc=True)
    dk = ts_utc.dt.tz_convert(TARGET_TZ)

    df = df.drop(columns=["timestamp"])
    df["dk"]        = dk
    df["dk_date"]   = dk.dt.date                 # dansk kalenderdato
    df["dk_hour"]   = dk.dt.hour                  # dansk time-paa-doegnet 0-23
    df["dk_dow"]    = dk.dt.dayofweek             # 0=mandag
    for c in ("open", "high", "low", "close", "volume"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(subset=["open", "high", "low", "close"])
    return df


# =============================================================================
# 2-3. BYG TIME-BARER + MAAL PR. TIME-BAR
# =============================================================================
def build_hour_bars(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Aggreger 1-min -> time-barer pr. (dansk dato, dansk time).

    Returnerer (gyldige_timebarer, tynde_timebarer). Tynde (n_min < graense)
    holdes separat saa de ikke forurener gennemsnittene.
    """
    # Path/churn: sum af |1-min afkast| (close-til-close) INDEN i hver time.
    # diff() inden for gruppen => foerste minut i timen bliver NaN (droppes af sum).
    df = df.sort_values("dk").copy()
    df["ret1"] = df.groupby(["dk_date", "dk_hour"])["close"].diff().abs()

    g = df.groupby(["dk_date", "dk_hour"], sort=True)
    hb = g.agg(
        o=("open", "first"),
        h=("high", "max"),
        l=("low", "min"),
        c=("close", "last"),
        vol=("volume", "sum"),
        n_min=("close", "size"),
        churn=("ret1", "sum"),
        dow=("dk_dow", "first"),
    ).reset_index()

    # De 6 maal (alle i indekspoint)
    hb["range"]    = hb["h"] - hb["l"]           # 1. raa bredde (retningsloes)
    hb["abs_move"] = (hb["c"] - hb["o"]).abs()   # 2. absolut netto-bevaegelse
    hb["net"]      = hb["c"] - hb["o"]           # 3. netto-afkast (fortegn)
    hb["max_up"]   = hb["h"] - hb["o"]           # 4. max op-traek fra open
    hb["max_down"] = hb["o"] - hb["l"]           # 5. max ned-traek fra open
    # 6. churn er allerede beregnet

    thin  = hb[hb["n_min"] < MIN_MIN_PER_HOUR].copy()
    valid = hb[hb["n_min"] >= MIN_MIN_PER_HOUR].copy()
    return valid, thin


# =============================================================================
# 4. "STORE TIMER" (konsistens)
# =============================================================================
def add_big_flag(valid: pd.DataFrame) -> pd.DataFrame:
    """Flag hver time-bar som 'stor' ift. dagens egen median-range (eller fast pt)."""
    valid = valid.copy()
    if BIG_HOUR_FIXED_PT is not None:
        thr = pd.Series(BIG_HOUR_FIXED_PT, index=valid.index)
    else:
        # dagens median-range paa tvaers af dagens timer (adaptiv skaering)
        thr = valid.groupby("dk_date")["range"].transform("median")
    valid["is_big"] = valid["range"] > thr
    return valid


# =============================================================================
# 4. AGGREGERING PR. DANSK TIME-PAA-DOEGNET
# =============================================================================
def bias_label(median_net: float) -> str:
    if median_net > BIAS_DEADZONE_PT:
        return "LONG"
    if median_net < -BIAS_DEADZONE_PT:
        return "SHORT"
    return "neutral"


def aggregate_by_hour(valid: pd.DataFrame) -> pd.DataFrame:
    """Median/gns/n + andel store + p90 + bias pr. dansk time (0-23)."""
    rows = []
    for hour in range(24):
        sub = valid[valid["dk_hour"] == hour]
        n = len(sub)
        row = {"dk_hour": hour, "n": n}
        for key in MEASURES:
            row[f"{key}_med"] = sub[key].median() if n else np.nan
            row[f"{key}_avg"] = sub[key].mean()   if n else np.nan
        # Volumen (kontrakter): skiller "bevaegelse MED likviditet" fra tynde spring.
        row["vol_med"]     = sub["vol"].median() if n else np.nan
        row["vol_avg"]     = sub["vol"].mean()   if n else np.nan
        row["range_p90"]   = sub["range"].quantile(0.90) if n else np.nan
        row["andel_store"] = sub["is_big"].mean() if n else np.nan
        row["bias"]        = bias_label(row["net_med"]) if n else "—"
        rows.append(row)
    return pd.DataFrame(rows)


# =============================================================================
# 5. TRAFIKLYS-VURDERING
# =============================================================================
def rel_terciles(agg: pd.DataFrame) -> dict[int, int]:
    """Relativ rang: del de gyldige timer (n>=MIN_N) i terciler efter median-range.
    Oeverste tredjedel -> 3 (groen-kandidat), midterste -> 2, nederste -> 1."""
    ok = agg[agg["n"] >= MIN_N].copy()
    if ok.empty:
        return {}
    ranks = ok["range_med"].rank(method="first")
    q = ranks / (len(ok) + 1e-9)
    level = {}
    for h, qq in zip(ok["dk_hour"], q):
        level[int(h)] = 3 if qq > 2/3 else (2 if qq > 1/3 else 1)
    return level


def assess_level(median_range: float, andel_store: float, n: float,
                 rel_level: int | None) -> int:
    """0=⚪ for lidt data, 1=🔴, 2=🟡, 3=🟢. Groen kraever BAADE absolut OG relativt.
    Konsistens (lav andel store timer) nedgraderer ét trin."""
    if not n or n < MIN_N or pd.isna(median_range):
        return 0
    # absolut
    if median_range >= ABS_GREEN_PT:
        abs_level = 3
    elif median_range >= ABS_YELLOW_PT:
        abs_level = 2
    else:
        abs_level = 1
    level = min(abs_level, rel_level if rel_level else 1)
    # konsistens-modifikator
    if andel_store is not None and not pd.isna(andel_store) \
            and andel_store < CONSISTENCY_MIN and level > 1:
        level -= 1
    return level


EMOJI = {0: "⚪", 1: "🔴", 2: "🟡", 3: "🟢"}


def level_text(level: int, andel_store: float, downgraded: bool) -> str:
    if level == 0:
        return "for lidt data / lukket"
    base = {1: "for stille", 2: "moderat", 3: "staerk"}[level]
    if level == 3:
        return "staerk, konsistent"
    if level == 2:
        return "moderat" + (", inkonsistent" if downgraded else "")
    return "for stille" + (", inkonsistent" if downgraded else "")


def add_assessment(agg: pd.DataFrame) -> pd.DataFrame:
    agg = agg.copy()
    rel = rel_terciles(agg)
    levels, texts, downgr = [], [], []
    for _, r in agg.iterrows():
        h = int(r["dk_hour"])
        rl = rel.get(h)
        lvl = assess_level(r["range_med"], r["andel_store"], r["n"], rl)
        # var den nedgraderet pga. konsistens?
        raw = assess_level(r["range_med"], 1.0, r["n"], rl)  # uden konsistens-straf
        dg = (lvl < raw)
        levels.append(lvl)
        downgr.append(dg)
        texts.append(f"{EMOJI[lvl]} {level_text(lvl, r['andel_store'], dg)}")
        _ = downgr
    agg["level"] = levels
    agg["vurdering"] = texts
    return agg


# =============================================================================
# 6. UGEDAG x TIME
# =============================================================================
def aggregate_by_dow_hour(valid: pd.DataFrame) -> pd.DataFrame:
    """Median-range + n + bias pr. (ugedag, time).

    Farve = RELATIV skala paa tvaers af ugens celler (terciler efter median-range):
    oeverste tredjedel groen, midterste gul, nederste roed. Relativ er det rette
    valg til et heatmap ('hvor i ugen ligger bevaegelsen'); den absolutte cost-floor
    (5/3 pt) klarer stort set alle MES-timer og ville farve alt groent.
    """
    rows = []
    for dow in sorted(valid["dow"].unique()):
        for hour in range(24):
            sub = valid[(valid["dow"] == dow) & (valid["dk_hour"] == hour)]
            n = len(sub)
            if n == 0:
                continue
            rows.append({
                "dow": int(dow), "ugedag": DK_DOW[int(dow)], "dk_hour": hour,
                "n": n, "range_med": sub["range"].median(),
                "abs_move_med": sub["abs_move"].median(),
                "vol_med": sub["vol"].median(),
                "net_med": sub["net"].median(), "bias": bias_label(sub["net"].median()),
            })
    dfc = pd.DataFrame(rows)
    if dfc.empty:
        dfc["level"] = []; dfc["vurdering"] = []
        return dfc

    # Relative terciler paa tvaers af cellerne med nok data
    ok = dfc[dfc["n"] >= MIN_N]
    if len(ok) >= 3:
        q1, q2 = ok["range_med"].quantile([1/3, 2/3])
    else:
        q1 = q2 = ok["range_med"].median() if len(ok) else 0.0

    def cell_level(r):
        if r["n"] < MIN_N:
            return 0
        if r["range_med"] > q2:
            return 3
        if r["range_med"] > q1:
            return 2
        return 1

    dfc["level"] = dfc.apply(cell_level, axis=1)
    dfc["vurdering"] = dfc["level"].map(EMOJI)
    return dfc


# =============================================================================
# 8. SPLIT-HALF ROBUSTHED (aar 1 vs aar 2)
# =============================================================================
def split_half_levels(valid: pd.DataFrame) -> pd.DataFrame:
    """Kør vurderingen paa foerste vs. anden halvdel af perioden; sammenlign."""
    dates = pd.to_datetime(pd.Series(sorted(valid["dk_date"].unique())))
    midpoint = dates.iloc[len(dates) // 2].date()

    def half_levels(mask):
        v = add_big_flag(valid[mask])
        a = add_assessment(aggregate_by_hour(v))
        return dict(zip(a["dk_hour"], a["level"]))

    m1 = valid["dk_date"] < midpoint
    m2 = valid["dk_date"] >= midpoint
    l1, l2 = half_levels(m1), half_levels(m2)

    rows = []
    for h in range(24):
        a, b = l1.get(h, 0), l2.get(h, 0)
        rows.append({
            "dk_hour": h, "level_h1": a, "level_h2": b,
            "robust": (a == b),
            "flag": ("groen kun i én halvdel" if (3 in (a, b) and a != b) else
                     ("skift" if a != b else "")),
        })
    return pd.DataFrame(rows), midpoint


# =============================================================================
# 7. OUTPUT — terminal, CSV, markdown
# =============================================================================
def hour_label(h: int) -> str:
    return f"{h:02d}:00-{(h+1) % 24:02d}:00"


def fmt(x, dec=1):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return "—"
    return f"{x:.{dec}f}"


def build_main_table(agg: pd.DataFrame) -> pd.DataFrame:
    """Hoveddeliverable-tabellen (24 raekker) med visningskolonner."""
    out = pd.DataFrame()
    out["Time (DK)"]        = agg["dk_hour"].map(hour_label)
    out["n (dage)"]         = agg["n"].astype(int)
    out["Median range (pt)"] = agg["range_med"].round(2)
    out["Range ($)"]        = (agg["range_med"] * USD_PER_POINT).round(0)
    out["Median |beveg.| (pt)"] = agg["abs_move_med"].round(2)
    out["Gns. afkast (pt,±)"]   = agg["net_avg"].round(2)
    out["Median max-op (pt)"]   = agg["max_up_med"].round(2)
    out["Median max-ned (pt)"]  = agg["max_down_med"].round(2)
    out["Churn (pt)"]           = agg["churn_med"].round(2)
    out["Median volumen"]       = agg["vol_med"].round(0)
    out["Range p90 (pt)"]       = agg["range_p90"].round(2)
    out["Andel store"]          = (agg["andel_store"] * 100).round(0)
    out["Bias"]                 = agg["bias"]
    out["Vurdering"]            = agg["vurdering"]
    return out


def print_terminal(main: pd.DataFrame, meta: dict):
    print("\n" + "=" * 100)
    print("  MES — BEVAEGELSE PR. TIME (DANSK TID)   |   ren beskrivende statistik, ingen strategi")
    print("=" * 100)
    print(f"  Data: {meta['rows']:,} 1-min bars  |  {meta['start']} -> {meta['slut']}  "
          f"|  {meta['days']} handelsdage")
    print(f"  Enhed: 1 pt = ${USD_PER_POINT:.0f}  |  median primaer, n vist  |  tz: {TARGET_TZ}")
    print("-" * 100)
    with pd.option_context("display.max_rows", None, "display.width", 200,
                           "display.max_columns", None):
        print(main.to_string(index=False))
    print("-" * 100)


def divergence_flags(agg: pd.DataFrame) -> list[str]:
    out = []
    for _, r in agg.iterrows():
        if r["n"] >= MIN_N and r["range_med"] and not pd.isna(r["range_med"]):
            if r["range_avg"] > r["range_med"] * DIVERGENCE_FACTOR:
                out.append(hour_label(int(r["dk_hour"])))
    return out


def best_worst(agg: pd.DataFrame) -> tuple[list, list]:
    ok = agg[agg["n"] >= MIN_N].sort_values("range_med", ascending=False)
    best = [(hour_label(int(r.dk_hour)), r.range_med, r.bias, r.vurdering)
            for r in ok.head(5).itertuples()]
    worst = [(hour_label(int(r.dk_hour)), r.range_med, r.vurdering)
             for r in ok.tail(5).itertuples()]
    return best, worst


def write_reports(main, agg, dow, split, midpoint, thin, meta):
    OUT_DIR.mkdir(exist_ok=True)

    # --- CSV 1: hoveddeliverable ---
    main_path = OUT_DIR / "handelstimer_dansktid.csv"
    main.to_csv(main_path, index=False, encoding="utf-8-sig")

    # --- CSV 2: ugedag x time (langt format) ---
    dow_path = OUT_DIR / "handelstimer_ugedag.csv"
    dow_out = dow.copy()
    dow_out["Time (DK)"] = dow_out["dk_hour"].map(hour_label)
    dow_out = dow_out[["ugedag", "Time (DK)", "dk_hour", "n", "range_med",
                       "abs_move_med", "vol_med", "net_med", "bias", "vurdering"]]
    dow_out.to_csv(dow_path, index=False, encoding="utf-8-sig")

    # --- MD-rapport ---
    best, worst = best_worst(agg)
    div = divergence_flags(agg)
    md = []
    md.append("# MES — handelstimer i dansk tid\n")
    md.append(f"*Ren beskrivende bevaegelses-statistik pr. time. Ingen strategi, ingen P&L.*\n")
    md.append(f"- **Data:** {meta['rows']:,} 1-min bars, {meta['start']} → {meta['slut']} "
              f"({meta['days']} handelsdage)\n")
    md.append(f"- **Enhed:** MES, 1 point = ${USD_PER_POINT:.0f} (1 tick = {POINT_PER_TICK} pt)\n")
    md.append(f"- **Tidszone:** alt i {TARGET_TZ} (dansk vaegur). Kilde er ET (tz-bevidst); "
              "konverteret med rigtig tz — dansk sommertid haandteret automatisk.\n")
    md.append("\n> **Note om sommertid:** USA og EU skifter sommertid paa forskellige "
              "datoer. I 2-3 uger om aaret daekker en 'dansk time' derfor et lidt andet "
              "markedsoejeblik end resten af aaret (fx US-aabningen rykker en dansk time "
              "i de uger). Det er korrekt: du handler efter dit eget vaegur, og det er "
              "praecis hvad tz-konverteringen giver.\n")

    md.append("\n## Bedste 3-5 timer at handle (dansk tid)\n")
    for lbl, mr, bias, vur in best:
        md.append(f"- **{lbl}** — median range {mr:.1f} pt (${mr*USD_PER_POINT:.0f}), "
                  f"bias {bias} — {vur}\n")
    md.append("\n## Undgaa typisk (mindst bevaegelse)\n")
    for lbl, mr, vur in worst:
        md.append(f"- {lbl} — median range {mr:.1f} pt — {vur}\n")

    md.append("\n## Alle 24 timer\n\n")
    md.append(main.to_markdown(index=False))
    md.append("\n")

    md.append("\n## Ugedag x time — median range (pt) + trafiklys\n\n")
    md.append(dow_matrix_md(dow))
    md.append("\n")

    md.append(f"\n## Split-half robusthed (foer/efter {midpoint})\n\n")
    md.append("Samme vurdering koert paa hver halvdel. En times 'groen' skal helst "
              "findes i begge halvdele for at vaere robust.\n\n")
    md.append(split_md(split))
    md.append("\n")

    if div:
        md.append("\n## Median/gennemsnit-divergens (faa ekstreme dage traekker gns. op)\n\n")
        md.append("Timer hvor gns-range > "
                  f"{DIVERGENCE_FACTOR}x median-range — brug medianen, ikke gns.: "
                  + ", ".join(div) + "\n")

    md.append("\n## Datakvalitet\n")
    md.append(f"- Tynde time-barer (<{MIN_MIN_PER_HOUR} min, typisk CME-vedligeholdelse "
              f"~23:00 DK + weekend-kanter): {len(thin):,} styk — holdt UDE af gennemsnittene.\n")
    md.append(f"- Timer med n < {MIN_N} dage vurderes ⚪ (for lidt data).\n")
    md.append("- CME: handel soendag aften → fredag; daglig vedligeholdelsespause "
              "17:00-18:00 ET (~23:00 DK). Weekend = lukket, ikke 'stille'.\n")

    (OUT_DIR / "handelstimer_rapport.md").write_text("".join(md), encoding="utf-8")
    return main_path, dow_path, OUT_DIR / "handelstimer_rapport.md"


def dow_matrix_md(dow: pd.DataFrame) -> str:
    """Matrix: raekker=timer 0-23, kolonner=ugedage. Celle = median range + emoji."""
    order = [d for d in range(7) if d in dow["dow"].values]
    cols = [DK_DOW[d] for d in order]
    lines = ["| Time (DK) | " + " | ".join(cols) + " |",
             "|---|" + "|".join(["---"] * len(cols)) + "|"]
    for h in range(24):
        cells = []
        for d in order:
            c = dow[(dow["dow"] == d) & (dow["dk_hour"] == h)]
            if c.empty:
                cells.append("—")
            else:
                r = c.iloc[0]
                cells.append(f"{EMOJI[int(r['level'])]} {r['range_med']:.1f} (n{int(r['n'])})")
        lines.append(f"| {hour_label(h)} | " + " | ".join(cells) + " |")
    return "\n".join(lines)


def split_md(split: pd.DataFrame) -> str:
    lines = ["| Time (DK) | 1. halvdel | 2. halvdel | Robust? | Flag |",
             "|---|---|---|---|---|"]
    for _, r in split.iterrows():
        lines.append(f"| {hour_label(int(r['dk_hour']))} | {EMOJI[int(r['level_h1'])]} "
                     f"| {EMOJI[int(r['level_h2'])]} | {'ja' if r['robust'] else 'NEJ'} "
                     f"| {r['flag']} |")
    return "\n".join(lines)


# =============================================================================
# EXCEL-UDGAVE (farvekodet heatmap)
# =============================================================================
# Excel-fyldfarver pr. trafiklys-niveau (Excels standard traffic-fills)
XLSX_FILL = {0: "D9D9D9", 1: "FFC7CE", 2: "FFEB9C", 3: "C6EFCE"}
XLSX_FONT = {0: "808080", 1: "9C0006", 2: "9C6500", 3: "006100"}


def write_excel(main, agg, dow, split, midpoint, meta):
    """Skriv en paen, farvekodet .xlsx med flere faner. Kraever openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OUT_DIR.mkdir(exist_ok=True)
    path = OUT_DIR / "handelstimer_mes.xlsx"

    wb = Workbook()
    thin_border = Border(*[Side(style="thin", color="DDDDDD")] * 4)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Fane 1: Timer (DK) — hoveddeliverable, farvet efter niveau ----
    ws = wb.active
    ws.title = "Timer (DK)"
    ws.append([f"MES — bevaegelse pr. time (dansk tid) · {meta['start']} → {meta['slut']} "
               f"· {meta['days']} handelsdage · 1 pt = ${USD_PER_POINT:.0f}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(main.columns))
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.append(list(main.columns))
    style_header(ws, len(main.columns), row=2)
    for i, (_, r) in enumerate(main.iterrows()):
        ws.append(list(r.values))
        lvl = int(agg.iloc[i]["level"])
        fill = PatternFill("solid", fgColor=XLSX_FILL[lvl])
        font = Font(color=XLSX_FONT[lvl], bold=(lvl == 3))
        rownum = ws.max_row
        for c in range(1, len(main.columns) + 1):
            cell = ws.cell(rownum, c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if c == len(main.columns):     # faerv kun "Vurdering"-cellen
                cell.fill = fill; cell.font = font
    ws.freeze_panes = "A3"
    autosize(ws, [12, 8, 16, 10, 18, 16, 15, 16, 11, 15, 13, 12, 9, 22])

    # ---- Feltforklaring UNDER hovedtabellen (hvert eneste felt beskrevet) ----
    # Eksplicit raekke-indeks (ikke ws.append) — tomme append-raekker desynkroniserer
    # ellers ws.max_row fra openpyxls interne markoer og overskriver felter.
    ncol = len(main.columns)
    row = 2 + len(main) + 2      # titel(1) + header(1) + datarraekker + 1 blank
    ws.cell(row, 1, "FELTFORKLARING — hvad hver kolonne betyder").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    row += 2

    def doc_block(title, items, row):
        """Skriv overskrift + raekker af (felt, beskrivelse). Returnerer naeste ledige raekke."""
        if title:
            ws.cell(row, 1, title).font = Font(bold=True, italic=True, color="1F3864")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
            row += 1
        for name, desc in items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c1 = ws.cell(row, 1, name)
            c1.font = Font(bold=True)
            c1.alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=ncol)
            c2 = ws.cell(row, 4, desc)
            c2.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = max(30, 15 * (1 + len(desc) // 95))
            row += 1
        return row + 1           # en blank raekke efter blokken

    up = f"{USD_PER_POINT:.0f}"
    field_docs = [
        ("Time (DK)",
         "Dansk time-interval i vaegur-tid (Europe/Copenhagen). Fx '15:00-16:00' = al "
         "bevaegelse mellem kl. 15:00:00 og 15:59:59 dansk tid, samlet paa tvaers af alle "
         "dage. Kilden er ET (US-boers) og konverteres tz-bevidst, saa dansk sommertid "
         "(CET<->CEST) haandteres automatisk."),
        ("n (dage)",
         "Antal dage der bidrager til timen (en observation pr. kalenderdato). Hoej n = "
         "mere paalideligt tal. Er n under 30 vurderes timen som 'for lidt data' (graa). "
         "Timer med faerre minutter end 30 (vedligeholdelse/weekend-kant) er sorteret fra "
         "foerst, saa de ikke traekker gennemsnittet."),
        ("Median range (pt)",
         "PRIMAERT bevaegelsesmaal. Median af (timens high - low) paa tvaers af dagene, i "
         "indekspoint. Det typiske spaend timen bevaeger sig i. Retningsloest (siger intet "
         "om op/ned). Median frem for gennemsnit, saa enkelte vilde nyhedsdage ikke oppuster "
         "tallet."),
        ("Range ($)",
         f"Median range omregnet til dollar: range_pt x ${up} (MES = ${up} pr. indekspoint). "
         "Hvad timens typiske spaend er vaerd i penge pr. kontrakt — den hurtige "
         "'kan det betale sig'-linje."),
        ("Median |beveg.| (pt)",
         "Median af |close - open|: hvor langt prisen faktisk endte fra sin start i timen "
         "(absolut netto, uden fortegn). Er typisk mindre end range, fordi prisen ofte "
         "vender undervejs. Stor forskel range vs. |beveg.| = timen churner (frem og "
         "tilbage) snarere end at trende."),
        ("Gns. afkast (pt, +/-)",
         "GENNEMSNIT af (close - open) MED fortegn. Positiv = timen tenderer op over de to "
         "aar, negativ = ned. Tal taet paa 0 = ingen retningsbias. Gennemsnit (ikke median) "
         "her, fordi vi vil fange systematisk drift, ikke den typiske dag."),
        ("Median max-op (pt)",
         "Median af (high - open): hvor langt OP timen typisk naaede fra sin aabning. "
         "Groft long-potentiale i timen — hvor meget en koeb-position kunne loebe med."),
        ("Median max-ned (pt)",
         "Median af (open - low): hvor langt NED timen typisk naaede fra sin aabning. "
         "Groft short-potentiale — hvor meget en salgs-position kunne loebe med."),
        ("Churn (pt)",
         "Median af summen af |1-minuts-afkast| inden i timen = total tilbagelagt vej "
         "(close-til-close). Hoej churn + LAV range = meget frem og tilbage om samme niveau "
         "(godt for mean-reversion). Hoej range + LAV churn = et enkelt rent spring."),
        ("Median volumen",
         "Median antal handlede kontrakter i timen (sum over minutterne, median over dagene). "
         "Skiller aegte bevaegelse MED likviditet fra tynde spring: en hoej range paa lav "
         "volumen er upaalidelig (svaer at komme ind/ud, bredere spread)."),
        ("Range p90 (pt)",
         "90-percentilen af range paa tvaers af dagene — 'haledagene'. Viser hvor stor timen "
         "kan blive paa de ~10% vildeste dage. Stor afstand fra medianen = timen har af og "
         "til eksplosive dage."),
        ("Andel store",
         "KONSISTENS-maal i procent: andel af dagene hvor timens range oversteg DAGENS EGEN "
         "median-range (paa tvaers af doegnets timer). Hoej = timen er paalideligt en af "
         "dagens mest aktive; lav = kun faa dage traekker den op. En stor-men-inkonsistent "
         "time nedgraderes et trin i vurderingen."),
        ("Bias",
         "Retningstendens ud fra median netto-afkast: LONG hvis klart positiv, SHORT hvis "
         "klart negativ, ellers 'neutral'. En lille doedzone (0,5 pt) forhindrer stoej i at "
         "blive kaldt en bias. NB: selv en 'bias' er svag paa MES — det er en tendens, "
         "ikke en regel."),
        ("Vurdering",
         "Trafiklys, samlet dom: GROEN = nok bevaegelse OG konsistent; GUL = moderat, vaelg "
         "dine spots; ROED = for stille/upaalidelig; GRAA = for lidt data eller lukket time. "
         "Kombinerer en absolut cost-floor (er der overhovedet nok?) med en relativ tercil "
         "(bedst vs. resten af doegnet) og nedgraderer inkonsistente timer."),
    ]
    row = doc_block("", field_docs, row)

    row = doc_block("Generelt", [
        ("Enhed",
         f"Alt i indekspoint (pt). 1 tick = {POINT_PER_TICK} pt; 1 pt = ${up} pr. MES-kontrakt. "
         "$-kolonnen er point x " + up + "."),
        ("Median vs. gennemsnit",
         "Median er primaer (robust mod enkelte ekstreme dage). Gennemsnit vises hvor det "
         "tilfoejer noget (afkast/drift). Stor forskel = faa dage traekker — stol paa medianen."),
        ("Ingen strategi",
         "Ren beskrivende historik. INGEN entry/exit, P&L, indikatorer eller forudsigelse. "
         "'Groen' betyder 'her er nok at arbejde med', IKKE 'her tjener man penge'."),
    ], row)

    row = doc_block("Andre faner i regnearket", [
        ("Ugedag x time (range)",
         "Matrix: raekker = de 24 danske timer, kolonner = ugedage. Celletal = median range "
         "(pt) for den ugedag+time. Farve = relativ tercil paa tvaers af ugens celler "
         "(groen=oeverste tredjedel, gul=midt, roed=nederste, graa=n<30)."),
        ("Ugedag x time (volumen)",
         "Samme matrix, men celletal = median volumen (kontrakter) og farve = relativ tercil "
         "af volumen. Læs sammen med range-matrixen: groen begge steder = bevaegelse MED "
         "likviditet."),
        ("Split-half",
         "Robusthedstjek. '1. halvdel'/'2. halvdel' = trafiklys for hver periode (foer/efter "
         "midtdatoen). 'Robust?' = samme farve i begge. 'Flag' = 'groen kun i en halvdel' "
         "(ikke robust) eller 'skift'. Et moenster der kun findes i det ene aar er ikke til "
         "at stole paa."),
        ("Raa aggregat",
         "Alle tal ufarvet til egen graving: dk_hour, n, <maal>_med og <maal>_avg for hvert "
         "af de 6 maal, vol_med/vol_avg, range_p90, andel_store, bias og level (0=graa, "
         "1=roed, 2=gul, 3=groen)."),
    ], row)

    # ---- Fane 2: Ugedag x time — heatmap (median range, farvet celle) ----
    ws2 = wb.create_sheet("Ugedag x time (range)")
    order = [d for d in range(7) if d in dow["dow"].values]
    cols = ["Time (DK)"] + [DK_DOW[d] for d in order]
    ws2.append(cols)
    style_header(ws2, len(cols))
    for h in range(24):
        rowvals = [hour_label(h)]
        levels = [None]
        for d in order:
            c = dow[(dow["dow"] == d) & (dow["dk_hour"] == h)]
            if c.empty:
                rowvals.append(None); levels.append(None)
            else:
                r = c.iloc[0]
                rowvals.append(round(float(r["range_med"]), 1))
                levels.append(int(r["level"]))
        ws2.append(rowvals)
        rownum = ws2.max_row
        ws2.cell(rownum, 1).font = Font(bold=True)
        ws2.cell(rownum, 1).border = thin_border
        for ci, lvl in enumerate(levels[1:], start=2):
            cell = ws2.cell(rownum, ci)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if lvl is not None:
                cell.fill = PatternFill("solid", fgColor=XLSX_FILL[lvl])
                cell.font = Font(color=XLSX_FONT[lvl])
    ws2.freeze_panes = "B2"
    autosize(ws2, [12] + [11] * len(order))
    ws2.append([])
    ws2.append(["Farve = relativ tercil paa tvaers af ugens celler (groen=top, gul=midt, roed=bund). "
                "Tal = median range i point. Graa = for lidt data (n<30)."])

    # ---- Fane 3: Ugedag x time — volumen (median kontrakter) ----
    ws3 = wb.create_sheet("Ugedag x time (volumen)")
    ws3.append(cols)
    style_header(ws3, len(cols))
    # relativ farvning af volumen paa tvaers af cellerne
    volcells = dow[dow["n"] >= MIN_N]["vol_med"]
    vq1, vq2 = (volcells.quantile([1/3, 2/3]) if len(volcells) >= 3
                else (volcells.median(), volcells.median())) if len(volcells) else (0, 0)
    for h in range(24):
        rowvals = [hour_label(h)]; vlevels = [None]
        for d in order:
            c = dow[(dow["dow"] == d) & (dow["dk_hour"] == h)]
            if c.empty:
                rowvals.append(None); vlevels.append(None)
            else:
                r = c.iloc[0]; v = float(r["vol_med"])
                rowvals.append(round(v))
                if r["n"] < MIN_N:
                    vlevels.append(0)
                else:
                    vlevels.append(3 if v > vq2 else (2 if v > vq1 else 1))
        ws3.append(rowvals)
        rownum = ws3.max_row
        ws3.cell(rownum, 1).font = Font(bold=True)
        for ci, lvl in enumerate(vlevels[1:], start=2):
            cell = ws3.cell(rownum, ci)
            cell.border = thin_border; cell.alignment = Alignment(horizontal="center")
            if lvl is not None:
                cell.fill = PatternFill("solid", fgColor=XLSX_FILL[lvl])
                cell.font = Font(color=XLSX_FONT[lvl])
    ws3.freeze_panes = "B2"
    autosize(ws3, [12] + [11] * len(order))
    ws3.append([])
    ws3.append(["Farve = relativ tercil af median-volumen (kontrakter) paa tvaers af ugens celler. "
                "Hoej range + hoej volumen = bevaegelse MED likviditet."])

    # ---- Fane 4: Split-half robusthed ----
    ws4 = wb.create_sheet("Split-half")
    ws4.append([f"Robusthed: samme vurdering paa foer/efter {midpoint}"])
    ws4.append(["Time (DK)", "1. halvdel", "2. halvdel", "Robust?", "Flag"])
    style_header(ws4, 5, row=2)
    for _, r in split.iterrows():
        ws4.append([hour_label(int(r["dk_hour"])),
                    EMOJI[int(r["level_h1"])], EMOJI[int(r["level_h2"])],
                    "ja" if r["robust"] else "NEJ", r["flag"]])
        rn = ws4.max_row
        for lvlcol, lvlval in ((2, int(r["level_h1"])), (3, int(r["level_h2"]))):
            cell = ws4.cell(rn, lvlcol)
            cell.fill = PatternFill("solid", fgColor=XLSX_FILL[lvlval])
            cell.alignment = Alignment(horizontal="center")
        if not r["robust"]:
            ws4.cell(rn, 4).font = Font(bold=True, color="9C0006")
    ws4.freeze_panes = "A3"
    autosize(ws4, [12, 12, 12, 10, 22])

    # ---- Fane 5: Raa aggregat (alle tal, ufarvet) ----
    ws5 = wb.create_sheet("Raa aggregat")
    raw_cols = ["dk_hour", "n"] + [f"{k}_med" for k in MEASURES] \
        + [f"{k}_avg" for k in MEASURES] + ["vol_med", "vol_avg", "range_p90",
                                            "andel_store", "bias", "level"]
    ws5.append(raw_cols)
    style_header(ws5, len(raw_cols))
    for _, r in agg.iterrows():
        ws5.append([round(r[c], 3) if isinstance(r[c], (int, float)) and not pd.isna(r[c])
                    else r[c] for c in raw_cols])
    ws5.freeze_panes = "A2"
    autosize(ws5, [9] * len(raw_cols))

    wb.save(path)
    return path


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"Laeser {DATA_FILE} ...")
    df = load_minute_data()
    meta = {
        "rows":  len(df),
        "start": str(df["dk"].min())[:16],
        "slut":  str(df["dk"].max())[:16],
        "days":  df["dk_date"].nunique(),
    }

    valid, thin = build_hour_bars(df)
    valid = add_big_flag(valid)

    agg = add_assessment(aggregate_by_hour(valid))
    main_tbl = build_main_table(agg)

    dow = aggregate_by_dow_hour(valid)
    split, midpoint = split_half_levels(valid)

    print_terminal(main_tbl, meta)

    best, worst = best_worst(agg)
    print("\n  BEDSTE TIMER (dansk tid):")
    for lbl, mr, bias, vur in best:
        print(f"    {lbl}  median range {mr:5.1f} pt (${mr*USD_PER_POINT:4.0f})  bias {bias:7s}  {vur}")
    print("\n  MINDST BEVAEGELSE:")
    for lbl, mr, vur in worst:
        print(f"    {lbl}  median range {mr:5.1f} pt  {vur}")

    n_robust = int(split["robust"].sum())
    print(f"\n  Split-half: {n_robust}/24 timer med samme trafiklys i begge halvdele "
          f"(skillelinje {midpoint}).")

    p1, p2, p3 = write_reports(main_tbl, agg, dow, split, midpoint, thin, meta)
    p4 = write_excel(main_tbl, agg, dow, split, midpoint, meta)
    print(f"\n  Skrevet:\n    {p1}\n    {p2}\n    {p3}\n    {p4}\n")


if __name__ == "__main__":
    main()
